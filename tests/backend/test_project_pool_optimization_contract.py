import base64
from io import BytesIO
from pathlib import Path
import unittest

import pandas as pd
from openpyxl import load_workbook
from pydantic import ValidationError

from app.schemas.workover_project_pool import (
    ImportTaskOut,
    WorkoverAttachment,
    WorkoverProjectPoolCreate,
)
from app.services.workover_project_pool_excel import (
    export_project_pool_template,
    parse_project_pool_excel,
)


REPO_ROOT = Path(__file__).resolve().parents[2]


class ProjectPoolOptimizationContractTest(unittest.TestCase):
    def test_project_pool_schema_exposes_business_quality_fields_and_attachments(self):
        fields = WorkoverProjectPoolCreate.model_fields

        for field_name in (
            "reason_category",
            "completeness_status",
            "data_source",
            "report_batch",
            "photo_requirement",
            "rejection_supplement",
            "is_duplicate_well",
            "related_project_ids",
            "attachments",
        ):
            self.assertIn(field_name, fields)

    def test_attachment_payload_carries_audit_and_business_metadata(self):
        attachment = WorkoverAttachment.model_validate(
            {
                "name": "井口照片.jpg",
                "url": "data:image/jpeg;base64,AA==",
                "content_type": "image/jpeg",
                "size": 128,
                "category": "wellsite",
                "uploaded_by": "张三",
                "uploaded_at": "2026-07-07T08:00:00Z",
            }
        )

        self.assertEqual(attachment.name, "井口照片.jpg")
        self.assertEqual(attachment.category, "wellsite")
        self.assertEqual(attachment.size, 128)

    def test_attachment_validation_rejects_invalid_file_metadata(self):
        with self.assertRaises(ValidationError):
            WorkoverAttachment.model_validate(
                {
                    "name": "",
                    "url": "not-a-url",
                    "content_type": "text/plain",
                    "size": -1,
                    "category": "",
                }
            )

    def test_excel_template_has_required_business_columns_and_instructions(self):
        template = export_project_pool_template()
        content = base64.b64decode(template["content_base64"])
        workbook = load_workbook(BytesIO(content))

        self.assertIn("项目池导入模板", workbook.sheetnames)
        self.assertIn("填写说明", workbook.sheetnames)
        headers = [cell.value for cell in workbook["项目池导入模板"][1]]

        for column in (
            "well_no",
            "report_unit",
            "reason_category",
            "data_source",
            "report_batch",
            "measure_type",
            "initiator_phone",
        ):
            self.assertIn(column, headers)

    def test_excel_import_reports_row_level_validation_errors(self):
        frame = pd.DataFrame(
            [
                {
                    "well_no": "CY2-001",
                    "report_unit": "采油一队",
                    "reason_category": "产量下降",
                    "measure_type": "pump_inspection",
                    "initiator_phone": "13800000001",
                },
                {
                    "well_no": "CY2-001",
                    "report_unit": "采油一队",
                    "reason_category": "产量下降",
                    "measure_type": "unknown_measure",
                    "initiator_phone": "bad-phone",
                },
            ]
        )
        output = BytesIO()
        frame.to_excel(output, index=False, engine="openpyxl")

        with self.assertRaises(ValueError) as ctx:
            parse_project_pool_excel(output.getvalue())

        message = str(ctx.exception)
        self.assertIn("第3行", message)
        self.assertIn("重复井号", message)
        self.assertIn("initiator_phone", message)
        self.assertIn("measure_type", message)

    def test_import_task_exposes_row_errors(self):
        self.assertIn("errors", ImportTaskOut.model_fields)

    def test_frontend_routes_separate_project_pool_ledger_from_approval_workbench(self):
        router_source = (REPO_ROOT / "frontend/src/router/index.ts").read_text(encoding="utf-8")
        project_view = REPO_ROOT / "frontend/src/views/ProjectPoolLedgerView.vue"
        approval_source = (REPO_ROOT / "frontend/src/views/ApprovalWorkbench.vue").read_text(encoding="utf-8")

        self.assertIn("ProjectPoolLedgerView", router_source)
        self.assertIn("/workover/project-pools", router_source)
        self.assertTrue(project_view.exists())
        self.assertNotIn("openCreate", approval_source)
        self.assertIn("审批处理", approval_source)

    def test_project_pool_ledger_exposes_detail_and_import_export_controls(self):
        source = (REPO_ROOT / "frontend/src/views/ProjectPoolLedgerView.vue").read_text(encoding="utf-8")

        for text in ("项目池台账", "导入", "下载模板", "导出", "详情", "资料完整性", "附件资料"):
            self.assertIn(text, source)

    def test_project_pool_and_approval_pages_do_not_duplicate_primary_work_surfaces(self):
        ledger_source = (REPO_ROOT / "frontend/src/views/ProjectPoolLedgerView.vue").read_text(encoding="utf-8")
        approval_source = (REPO_ROOT / "frontend/src/views/ApprovalWorkbench.vue").read_text(encoding="utf-8")

        self.assertIn("source-summary", ledger_source)
        self.assertNotIn("workflow-strip", ledger_source)
        self.assertNotIn("地质核实", ledger_source)
        self.assertNotIn("工艺核实", ledger_source)

        self.assertIn("workflow-strip", approval_source)
        self.assertNotIn("el-drawer", approval_source)
        self.assertNotIn("openDetail", approval_source)
        self.assertNotIn("单井详情", approval_source)

    def test_core_documents_and_views_keep_readable_chinese(self):
        readme = (REPO_ROOT / "README.md").read_text(encoding="utf-8")
        model_source = (REPO_ROOT / "app/models/workover.py").read_text(encoding="utf-8")

        self.assertIn("采油二厂井下作业管理系统", readme)
        self.assertIn("上修项目池表", model_source)
        self.assertNotIn("閲", readme)
        self.assertNotIn("椤圭", model_source)


if __name__ == "__main__":
    unittest.main()
