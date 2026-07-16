"""A5 数据同步服务。

包含日报同步、异常同步、全量同步的核心业务逻辑。
每 30 分钟由 Celery 定时任务触发。
"""

import logging
import base64
import hashlib
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

import httpx
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import BusinessException
from app.core.redis import cache_client
from app.core.status_codes import A5_LINK_FAILED, CONFLICT
from app.models.integration import IntegrationEvent, IntegrationEventStatus
from app.models.workover import ContractorCapacityStatus, OperationStatus, ProjectPoolStatus, WorkoverOperationSheet
from app.schemas.a5_integration import A5AnalyticsOut, A5AnalyticsQuery, A5NameValueOut, A5TrendOut, A5AnalyticsReportOut
from app.services.a5_adapter import get_a5_adapter
from app.services.a5_data_cleaner import clean_daily_report, validate_operation_data

logger = logging.getLogger(__name__)

A5_SYNC_STATUS_KEY = "a5:sync:last_status"
A5_SYNC_COUNT_PREFIX = "a5:sync:count:"
A5_ANOMALY_RECORDS_KEY = "a5:sync:anomaly_records"
A5_PROCESS_RECORDS_KEY = "a5:sync:process_records"
A5_ANOMALY_RECORDS_PREFIX = f"{A5_ANOMALY_RECORDS_KEY}:"
A5_PROCESS_RECORDS_PREFIX = f"{A5_PROCESS_RECORDS_KEY}:"
A5_ANOMALY_DATES_KEY = "a5:sync:anomaly_dates"
A5_PROCESS_DATES_KEY = "a5:sync:process_dates"
A5_ANALYTICS_CACHE_TTL = 604800


@dataclass(frozen=True)
class A5EventProcessResult:
    event: IntegrationEvent
    duplicate: bool
    matched: bool


def process_a5_callback_event(
    db: Session,
    payload: dict[str, Any],
    *,
    event_id: str | None = None,
) -> A5EventProcessResult:
    canonical_payload = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    payload_hash = hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()
    event_key = event_id.strip() if event_id and event_id.strip() else payload_hash

    existing = db.scalar(
        select(IntegrationEvent).where(
            IntegrationEvent.source == "a5",
            IntegrationEvent.event_key == event_key,
        )
    )
    if existing is not None:
        if existing.payload_hash != payload_hash:
            raise BusinessException(CONFLICT, "A5 事件键与既有载荷冲突")
        return A5EventProcessResult(
            event=existing,
            duplicate=True,
            matched=existing.status == IntegrationEventStatus.PROCESSED,
        )

    operation_no = str(payload.get("operation_no") or "")
    event = IntegrationEvent(
        source="a5",
        event_key=event_key,
        payload_hash=payload_hash,
        operation_no=operation_no or None,
        raw_payload=payload,
        attempt_count=1,
    )
    try:
        # The unique key is the concurrency control.  A savepoint keeps the
        # surrounding session usable when two callbacks arrive at once.
        with db.begin_nested():
            db.add(event)
            db.flush()
    except IntegrityError:
        db.expire_all()
        existing = db.scalar(
            select(IntegrationEvent).where(
                IntegrationEvent.source == "a5",
                IntegrationEvent.event_key == event_key,
            )
        )
        if existing is None:
            raise
        if existing.payload_hash != payload_hash:
            raise BusinessException(CONFLICT, "A5 事件键与既有载荷冲突")
        return A5EventProcessResult(
            event=existing,
            duplicate=True,
            matched=existing.status == IntegrationEventStatus.PROCESSED,
        )

    sheet = db.scalar(
        select(WorkoverOperationSheet).where(WorkoverOperationSheet.operation_no == operation_no)
    )
    if sheet is None:
        event.status = IntegrationEventStatus.PENDING_REVIEW
        event.error_message = "未找到对应作业工单"
        db.commit()
        db.refresh(event)
        return A5EventProcessResult(event=event, duplicate=False, matched=False)

    apply_a5_update_to_operation_sheet(
        sheet,
        status=str(payload.get("status") or ""),
        remark=payload.get("remark"),
        detail=payload,
        source="callback",
    )
    event.status = IntegrationEventStatus.PROCESSED
    event.processed_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(event)
    return A5EventProcessResult(event=event, duplicate=False, matched=True)


def _normalize_a5_status(raw_status: str | None) -> OperationStatus | None:
    if not raw_status:
        return None
    status = raw_status.strip().upper()
    status_map = {
        "通过": OperationStatus.FINISHED,
        "办结": OperationStatus.FINISHED,
        "完成": OperationStatus.FINISHED,
        "已完成": OperationStatus.FINISHED,
        "FINISHED": OperationStatus.FINISHED,
        "DONE": OperationStatus.FINISHED,
        "驳回": OperationStatus.WAITING_DISPATCH,
        "退回": OperationStatus.WAITING_DISPATCH,
        "REJECTED": OperationStatus.WAITING_DISPATCH,
        "关闭": OperationStatus.CANCELED,
        "取消": OperationStatus.CANCELED,
        "CANCELED": OperationStatus.CANCELED,
        "CANCELLED": OperationStatus.CANCELED,
        "开工": OperationStatus.WORKING,
        "施工中": OperationStatus.WORKING,
        "作业中": OperationStatus.WORKING,
        "WORKING": OperationStatus.WORKING,
        "下发": OperationStatus.DISPATCHED,
        "已下发": OperationStatus.DISPATCHED,
        "审核中": OperationStatus.DISPATCHED,
        "DISPATCHED": OperationStatus.DISPATCHED,
    }
    return status_map.get(status)


def _normalize_progress(raw_progress: Any) -> int | None:
    if raw_progress is None or raw_progress == "":
        return None
    if isinstance(raw_progress, str):
        raw_progress = raw_progress.strip().rstrip("%")
    try:
        return max(0, min(100, int(float(raw_progress))))
    except (TypeError, ValueError):
        return None


OCCUPIED_OPERATION_STATUSES = {OperationStatus.DISPATCHED, OperationStatus.WORKING}


def _cache_a5_records(
    *,
    latest_key: str,
    prefix: str,
    dates_key: str,
    sync_date: str,
    records: list[dict[str, Any]],
) -> None:
    cache_client.set_json(f"{prefix}{sync_date}", records, expire_seconds=A5_ANALYTICS_CACHE_TTL)
    cache_client.set_json(latest_key, records, expire_seconds=A5_ANALYTICS_CACHE_TTL)

    raw_dates = cache_client.get_json(dates_key) or []
    dates = {str(item) for item in raw_dates if item}
    dates.add(sync_date)
    cache_client.set_json(dates_key, sorted(dates)[-31:], expire_seconds=A5_ANALYTICS_CACHE_TTL)


def _date_in_query_range(day_text: str, query: A5AnalyticsQuery) -> bool:
    try:
        day = datetime.strptime(day_text, "%Y-%m-%d").date()
    except ValueError:
        return False
    if query.start_date and day < query.start_date:
        return False
    if query.end_date and day > query.end_date:
        return False
    return True


def _load_cached_a5_records(
    *,
    latest_key: str,
    prefix: str,
    dates_key: str,
    query: A5AnalyticsQuery,
) -> list[dict[str, Any]]:
    raw_dates = cache_client.get_json(dates_key) or []
    dates = sorted({str(item) for item in raw_dates if item and _date_in_query_range(str(item), query)})

    records: list[dict[str, Any]] = []
    for day_text in dates:
        day_records = cache_client.get_json(f"{prefix}{day_text}") or []
        if isinstance(day_records, list):
            records.extend(item for item in day_records if isinstance(item, dict))

    if not records and not raw_dates:
        legacy_records = cache_client.get_json(latest_key) or []
        if isinstance(legacy_records, list):
            records.extend(item for item in legacy_records if isinstance(item, dict))
    return records


def apply_a5_update_to_operation_sheet(
    sheet: WorkoverOperationSheet,
    *,
    status: str | None,
    remark: str | None = None,
    progress: int | None = None,
    detail: dict[str, Any] | None = None,
    source: str,
) -> OperationStatus | None:
    """Apply A5 status/report data to an operation sheet in one consistent place."""
    now = datetime.now(timezone.utc)
    new_status = _normalize_a5_status(status)
    old_status = sheet.status

    sheet.a5_status = status
    sheet.a5_remark = remark
    sheet.last_a5_sync_at = now
    merged_detail = dict(sheet.progress_detail or {})
    merged_detail[f"a5_{source}"] = {
        "status": status,
        "remark": remark,
        "synced_at": now.isoformat(),
        "raw": detail or {},
    }
    sheet.progress_detail = merged_detail

    normalized_progress = _normalize_progress(progress)
    if normalized_progress is not None:
        sheet.progress = normalized_progress

    if new_status is None:
        return None

    contractor = sheet.contractor_capacity
    sheet.status = new_status
    if new_status == OperationStatus.DISPATCHED and sheet.progress < 1:
        sheet.progress = 1
    elif new_status == OperationStatus.WORKING:
        if sheet.progress < 1:
            sheet.progress = 1
        if sheet.actual_start_at is None:
            sheet.actual_start_at = now
    elif new_status == OperationStatus.FINISHED:
        sheet.progress = 100
        if sheet.actual_start_at is None:
            sheet.actual_start_at = now
        sheet.actual_end_at = now
        if contractor is not None and old_status in OCCUPIED_OPERATION_STATUSES:
            contractor.available_count += 1
            contractor.status = ContractorCapacityStatus.AVAILABLE
    elif new_status in {OperationStatus.WAITING_DISPATCH, OperationStatus.CANCELED}:
        if new_status == OperationStatus.WAITING_DISPATCH:
            sheet.contractor_capacity_id = None
            sheet.progress = 0
            if sheet.project is not None:
                sheet.project.status = ProjectPoolStatus.APPROVED
        if contractor is not None and old_status in OCCUPIED_OPERATION_STATUSES:
            contractor.available_count += 1
            contractor.status = ContractorCapacityStatus.AVAILABLE

    return new_status


def build_local_daily_reports(
    sheets: list[WorkoverOperationSheet],
    sync_date: str,
) -> list[dict[str, Any]]:
    reports: list[dict[str, Any]] = []
    for sheet in sheets:
        if sheet.status == OperationStatus.DISPATCHED:
            reports.append(
                {
                    "operation_no": sheet.operation_no,
                    "status": "WORKING",
                    "progress": max(sheet.progress or 0, 35),
                    "report_date": sync_date,
                    "remark": "本地演示模拟：A5 已接收派工并进入施工",
                }
            )
        elif sheet.status == OperationStatus.WORKING:
            next_progress = 100 if (sheet.progress or 0) >= 80 else max(sheet.progress or 0, 65)
            reports.append(
                {
                    "operation_no": sheet.operation_no,
                    "status": "FINISHED" if next_progress == 100 else "WORKING",
                    "progress": next_progress,
                    "report_date": sync_date,
                    "remark": "本地演示模拟：A5 日报同步更新施工进度",
                }
            )
    return reports


async def sync_daily_operations(db: Session, sync_date: str | None = None) -> dict[str, Any]:
    """同步 A5 日报数据到修井运行表。

    Args:
        db: 数据库会话
        sync_date: 同步日期（YYYY-MM-DD），默认当天

    Returns:
        同步统计：{total, updated, failed}
    """
    if sync_date is None:
        sync_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if settings.a5_adapter_mode == "mock":
        raw_data = build_local_daily_reports(
            list(
                db.query(WorkoverOperationSheet)
                .filter(WorkoverOperationSheet.status.in_([OperationStatus.DISPATCHED, OperationStatus.WORKING]))
                .all()
            ),
            sync_date,
        )
    else:
        try:
            raw_data = await get_a5_adapter().fetch_daily_reports(date.fromisoformat(sync_date))
        except BusinessException as exc:
            logger.error(f"A5 日报拉取失败: {exc.msg}")
            return {"total": 0, "updated": 0, "failed": 0, "error": exc.msg}

    cleaned = clean_daily_report(raw_data)
    stats = {"total": len(cleaned), "updated": 0, "failed": 0}

    for record in cleaned:
        if not validate_operation_data(record):
            stats["failed"] += 1
            continue
        try:
            operation_no = record.get("operation_no", "")
            existing = db.query(WorkoverOperationSheet).filter(
                WorkoverOperationSheet.operation_no == operation_no
            ).first()
            if existing:
                apply_a5_update_to_operation_sheet(
                    existing,
                    status=record.get("status") or record.get("operation_status") or "施工中",
                    remark=record.get("remark"),
                    progress=record.get("progress"),
                    detail=record,
                    source="daily_report",
                )
                stats["updated"] += 1
        except Exception as exc:
            logger.exception(f"日报同步更新失败: {exc}")
            db.rollback()
            stats["failed"] += 1

    if stats["updated"] > 0 or stats["total"] > 0:
        db.commit()

    logger.info(f"A5 日报同步完成: {stats}")
    return stats


async def sync_anomalies(db: Session, sync_date: str | None = None) -> dict[str, Any]:
    """同步 A5 施工异常数据。"""
    if sync_date is None:
        sync_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    try:
        raw_data = await get_a5_adapter().fetch_anomalies(date.fromisoformat(sync_date))
    except BusinessException as exc:
        logger.error(f"A5 异常数据拉取失败: {exc.msg}")
        return {"total": 0, "synced": 0, "error": exc.msg}

    cleaned = clean_daily_report(raw_data)
    _cache_a5_records(
        latest_key=A5_ANOMALY_RECORDS_KEY,
        prefix=A5_ANOMALY_RECORDS_PREFIX,
        dates_key=A5_ANOMALY_DATES_KEY,
        sync_date=sync_date,
        records=cleaned,
    )
    logger.info(f"A5 异常同步完成: {len(cleaned)} 条")
    return {"total": len(cleaned), "synced": len(cleaned)}


async def sync_process_progress(db: Session, sync_date: str | None = None) -> dict[str, Any]:
    """同步 A5 特殊工序/工序进度数据。"""
    if sync_date is None:
        sync_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    try:
        raw_data = await get_a5_adapter().fetch_process_progress(date.fromisoformat(sync_date))
    except BusinessException as exc:
        logger.error(f"A5 工序数据拉取失败: {exc.msg}")
        return {"total": 0, "synced": 0, "error": exc.msg}

    cleaned = clean_daily_report(raw_data)
    _cache_a5_records(
        latest_key=A5_PROCESS_RECORDS_KEY,
        prefix=A5_PROCESS_RECORDS_PREFIX,
        dates_key=A5_PROCESS_DATES_KEY,
        sync_date=sync_date,
        records=cleaned,
    )
    logger.info(f"A5 工序同步完成: {len(cleaned)} 条")
    return {"total": len(cleaned), "synced": len(cleaned)}


def _record_day(record: dict[str, Any]) -> date | None:
    raw = (
        record.get("date")
        or record.get("report_date")
        or record.get("created_at")
        or record.get("time")
        or record.get("occurred_at")
    )
    if not raw:
        return None
    try:
        return datetime.fromisoformat(str(raw).replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return datetime.strptime(str(raw)[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def _record_category(record: dict[str, Any], keys: tuple[str, ...], fallback: str) -> str:
    for key in keys:
        value = record.get(key)
        if value:
            return str(value)
    return fallback


def _filter_records(records: list[dict[str, Any]], query: A5AnalyticsQuery, keys: tuple[str, ...]) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for record in records:
        day = _record_day(record)
        if (query.start_date or query.end_date) and day is None:
            continue
        if query.start_date and day and day < query.start_date:
            continue
        if query.end_date and day and day > query.end_date:
            continue
        if query.category and query.category not in _record_category(record, keys, ""):
            continue
        filtered.append(record)
    return filtered


def build_a5_analytics(query: A5AnalyticsQuery) -> A5AnalyticsOut:
    """Build cached A5 anomaly/special-process statistics for charts and reports."""
    anomaly_records = _load_cached_a5_records(
        latest_key=A5_ANOMALY_RECORDS_KEY,
        prefix=A5_ANOMALY_RECORDS_PREFIX,
        dates_key=A5_ANOMALY_DATES_KEY,
        query=query,
    )
    process_records = _load_cached_a5_records(
        latest_key=A5_PROCESS_RECORDS_KEY,
        prefix=A5_PROCESS_RECORDS_PREFIX,
        dates_key=A5_PROCESS_DATES_KEY,
        query=query,
    )
    anomaly_keys = ("anomaly_type", "exception_type", "type", "category")
    process_keys = ("process_type", "procedure_type", "measure_type", "category")

    anomalies = _filter_records(anomaly_records, query, anomaly_keys)
    processes = _filter_records(process_records, query, process_keys)

    anomaly_counter = Counter(_record_category(item, anomaly_keys, "未分类异常") for item in anomalies)
    process_counter = Counter(_record_category(item, process_keys, "未分类工序") for item in processes)

    trend_map: dict[str, dict[str, int]] = defaultdict(lambda: {"anomaly": 0, "process": 0})
    for item in anomalies:
        day = _record_day(item)
        if day:
            trend_map[day.isoformat()]["anomaly"] += 1
    for item in processes:
        day = _record_day(item)
        if day:
            trend_map[day.isoformat()]["process"] += 1

    days = sorted(trend_map)
    return A5AnalyticsOut(
        anomaly_total=len(anomalies),
        special_process_total=len(processes),
        anomaly_distribution=[A5NameValueOut(name=name, value=count) for name, count in anomaly_counter.most_common()],
        process_distribution=[A5NameValueOut(name=name, value=count) for name, count in process_counter.most_common()],
        trend=A5TrendOut(
            days=days,
            anomaly_counts=[trend_map[day]["anomaly"] for day in days],
            process_counts=[trend_map[day]["process"] for day in days],
        ),
    )


def export_a5_analytics_report(query: A5AnalyticsQuery, template_name: str | None = None) -> A5AnalyticsReportOut:
    """Export A5 anomaly and special-process statistics as an Excel report."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    summary = build_a5_analytics(query)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A5统计报告"

    title = "A5异常与特殊工序统计报告"
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["统计起始", query.start_date.isoformat() if query.start_date else "不限", "统计截止", query.end_date.isoformat() if query.end_date else "不限"])
    ws.append(["类别条件", query.category or "全部", "模板", template_name or "默认统计模板"])
    ws.append(["异常总数", summary.anomaly_total, "特殊工序总数", summary.special_process_total])
    ws.append([])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws.append(["异常类别", "数量", "特殊工序类别", "数量"])
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    max_rows = max(len(summary.anomaly_distribution), len(summary.process_distribution))
    for index in range(max_rows):
        anomaly = summary.anomaly_distribution[index] if index < len(summary.anomaly_distribution) else None
        process = summary.process_distribution[index] if index < len(summary.process_distribution) else None
        ws.append([
            anomaly.name if anomaly else "",
            anomaly.value if anomaly else "",
            process.name if process else "",
            process.value if process else "",
        ])

    trend_start = ws.max_row + 2
    ws.cell(row=trend_start, column=1, value="日期")
    ws.cell(row=trend_start, column=2, value="异常数量")
    ws.cell(row=trend_start, column=3, value="特殊工序数量")
    for cell in ws[trend_start]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for day, anomaly_count, process_count in zip(
        summary.trend.days,
        summary.trend.anomaly_counts,
        summary.trend.process_counts,
    ):
        ws.append([day, anomaly_count, process_count])

    for column in ("A", "B", "C", "D"):
        ws.column_dimensions[column].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"a5_analytics_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
    return A5AnalyticsReportOut(
        filename=filename,
        content_base64=base64.b64encode(output.read()).decode("ascii"),
    )


async def full_sync(db: Session) -> dict[str, Any]:
    """全量同步 - 组合日报 + 异常 + 工序进度。"""
    sync_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    daily_stats = await sync_daily_operations(db, sync_date)
    anomaly_stats = await sync_anomalies(db, sync_date)
    process_stats = await sync_process_progress(db, sync_date)

    result = {
        "sync_time": datetime.now(timezone.utc).isoformat(),
        "daily": daily_stats,
        "anomaly": anomaly_stats,
        "process": process_stats,
        "overall": "success",
    }

    error_message = daily_stats.get("error") or anomaly_stats.get("error") or process_stats.get("error")
    if error_message:
        result["overall"] = "partial_failure"

    today_key = f"{A5_SYNC_COUNT_PREFIX}{sync_date}"
    count = cache_client.get_json(today_key) or 0
    cache_client.set_json(today_key, int(count) + 1, expire_seconds=172800)
    cache_client.set_json(
        A5_SYNC_STATUS_KEY,
        {
            "last_sync_time": result["sync_time"],
            "last_sync_status": result["overall"],
            "last_sync_message": (
                error_message
                or (
                    f"日报更新 {daily_stats.get('updated', 0)} 条，"
                    f"异常同步 {anomaly_stats.get('synced', 0)} 条，"
                    f"工序同步 {process_stats.get('synced', 0)} 条"
                )
            ),
            "sync_count_today": int(count) + 1,
            "is_running": False,
        },
        expire_seconds=604800,
    )
    if error_message:
        raise BusinessException(A5_LINK_FAILED, str(error_message))
    return result


def _trigger_alert(message: str) -> None:
    """触发企业微信/内部通告警。"""
    logger.error(f"[A5 告警] {message}")
    if not settings.alert_webhook_url:
        return
    try:
        with httpx.Client(timeout=5) as client:
            client.post(
                settings.alert_webhook_url,
                json={"msgtype": "text", "text": {"content": message}},
            )
    except Exception:
        logger.exception("告警 Webhook 推送失败")
